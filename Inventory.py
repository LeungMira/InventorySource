import tkinter as tk
from tkinter import messagebox, filedialog, ttk
import openpyxl
from datetime import datetime

# Global variable to store the current file path
current_file_path = None

# Functionality for adding items
def add_item():
    item = entry_item.get()
    quantity = entry_quantity.get()
    price = entry_price.get()

    if item and quantity and price:
        try:
            quantity = int(quantity)
            price = float(price)
            merge_items(item, quantity, price)
            entry_item.delete(0, tk.END)
            entry_quantity.delete(0, tk.END)
            entry_price.delete(0, tk.END)
        except ValueError:
            messagebox.showerror("Invalid Input", "Quantity must be a number and price must be a valid number.")
    else:
        messagebox.showerror("Error", "All fields must be filled!")

# Functionality for merging duplicate items
def merge_items(new_item, new_quantity, new_price):
    found = False
    for i in range(listbox.size()):
        entry = listbox.get(i)
        existing_item, existing_quantity, existing_price = entry.split(" - ")
        existing_quantity = int(existing_quantity)

        if existing_item.lower() == new_item.lower():
            found = True
            updated_quantity = existing_quantity + new_quantity
            listbox.delete(i)
            listbox.insert(i, f"{existing_item} - {updated_quantity} - {new_price:.2f}")
            break

    if not found:
        listbox.insert(tk.END, f"{new_item} - {new_quantity} - {new_price:.2f}")

# Functionality for saving items to Excel with preserved history
def save_to_excel():
    global current_file_path
    if not current_file_path:
        current_file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not current_file_path:
            messagebox.showerror("Error", "No file selected or created.")
            return

    try:
        if current_file_path:
            try:
                workbook = openpyxl.load_workbook(current_file_path)
            except FileNotFoundError:
                workbook = openpyxl.Workbook()

            if "Current Inventory" in workbook.sheetnames:
                inventory_sheet = workbook["Current Inventory"]
                inventory_sheet.delete_rows(2, inventory_sheet.max_row)
            else:
                inventory_sheet = workbook.create_sheet(title="Current Inventory")
                inventory_sheet.append(["Item", "Quantity", "Price per Unit"])

            for entry in listbox.get(0, tk.END):
                item, quantity, price = entry.split(" - ")
                inventory_sheet.append([item, quantity, float(price)])

            if "Consumption History" not in workbook.sheetnames:
                history_sheet = workbook.create_sheet(title="Consumption History")
                history_sheet.append(["Action", "Item", "Quantity", "Consumer", "Date", "Price per Unit"])

            workbook.save(current_file_path)
            messagebox.showinfo("Success", "Data saved successfully!")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to save data: {e}")
        current_file_path = None  # Reset file path if save fails

# Functionality for loading items from Excel
def load_from_excel():
    global current_file_path
    current_file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if current_file_path:
        try:
            workbook = openpyxl.load_workbook(current_file_path)
            inventory_sheet = workbook["Current Inventory"]

            listbox.delete(0, tk.END)
            for row in inventory_sheet.iter_rows(min_row=2, values_only=True):
                item, quantity, price = row
                listbox.insert(tk.END, f"{item} - {quantity} - {float(price):.2f}")

            messagebox.showinfo("Success", "Data loaded successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load data: {e}")

# Functionality for consumption (delete or edit items) and record updates
def consumption():
    item = entry_consume_item.get()
    quantity = entry_consume_quantity.get()
    consumer = entry_consumer.get()

    if not (item and quantity and consumer):
        messagebox.showerror("Error", "All fields must be filled for consumption!")
        return

    try:
        quantity = int(quantity)
        found = False

        if not current_file_path:
            messagebox.showerror("Error", "No file loaded. Please load a file first.")
            return

        workbook = openpyxl.load_workbook(current_file_path)
        inventory_sheet = workbook["Current Inventory"]
        history_sheet = workbook["Consumption History"]

        for i in range(listbox.size()):
            entry = listbox.get(i)
            existing_item, existing_quantity, existing_price = entry.split(" - ")
            existing_quantity = int(existing_quantity)

            if existing_item.lower() == item.lower():
                found = True
                current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                if existing_quantity > quantity:
                    new_quantity = existing_quantity - quantity
                    listbox.delete(i)
                    listbox.insert(i, f"{existing_item} - {new_quantity} - {existing_price}")
                    inventory_sheet.cell(row=i + 2, column=2, value=new_quantity)
                    history_sheet.append(["Consumed", existing_item, quantity, consumer, current_date, float(existing_price)])
                    messagebox.showinfo("Consumption Recorded", f"{consumer} consumed {quantity} of {existing_item}.")
                elif existing_quantity == quantity:
                    listbox.delete(i)
                    inventory_sheet.delete_rows(i + 2)
                    history_sheet.append(["Consumed", existing_item, quantity, consumer, current_date, float(existing_price)])
                    messagebox.showinfo("Consumption Recorded", f"{consumer} consumed and depleted {existing_item}.")
                else:
                    messagebox.showerror("Error", "Consumption exceeds available quantity.")
                break

        if not found:
            messagebox.showerror("Error", f"Item '{item}' not found in inventory.")

        entry_consume_item.delete(0, tk.END)
        entry_consume_quantity.delete(0, tk.END)
        entry_consumer.delete(0, tk.END)

        workbook.save(current_file_path)

    except ValueError:
        messagebox.showerror("Invalid Input", "Quantity must be a number.")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to record consumption: {e}")

# Main application window
root = tk.Tk()
root.title("Inventory Manager")

notebook = ttk.Notebook(root)
notebook.pack(expand=True, fill="both")

# Add Inventory Tab
frame_add = ttk.Frame(notebook)
notebook.add(frame_add, text="Add Inventory")

label_item = tk.Label(frame_add, text="Item:")
label_item.grid(row=0, column=0, padx=10, pady=10)

entry_item = tk.Entry(frame_add)
entry_item.grid(row=0, column=1, padx=10, pady=10)

label_quantity = tk.Label(frame_add, text="Quantity:")
label_quantity.grid(row=1, column=0, padx=10, pady=10)

entry_quantity = tk.Entry(frame_add)
entry_quantity.grid(row=1, column=1, padx=10, pady=10)

label_price = tk.Label(frame_add, text="Price per Unit:")
label_price.grid(row=2, column=0, padx=10, pady=10)

entry_price = tk.Entry(frame_add)
entry_price.grid(row=2, column=1, padx=10, pady=10)

btn_add = tk.Button(frame_add, text="Add Item", command=add_item)
btn_add.grid(row=3, column=0, columnspan=2, pady=10)

# Inventory Display Tab
frame_display = ttk.Frame(notebook)
notebook.add(frame_display, text="View Inventory")

listbox = tk.Listbox(frame_display, width=60)
listbox.grid(row=0, column=0, columnspan=2, padx=10, pady=10)

btn_save = tk.Button(frame_display, text="Save to Excel", command=save_to_excel)
btn_save.grid(row=1, column=0, padx=10, pady=10)

btn_load = tk.Button(frame_display, text="Load from Excel", command=load_from_excel)
btn_load.grid(row=1, column=1, padx=10, pady=10)

# Consumption Tab
frame_consume = ttk.Frame(notebook)
notebook.add(frame_consume, text="Consumption")

label_consume_item = tk.Label(frame_consume, text="Item:")
label_consume_item.grid(row=0, column=0, padx=10, pady=10)

entry_consume_item = tk.Entry(frame_consume)
entry_consume_item.grid(row=0, column=1, padx=10, pady=10)

label_consume_quantity = tk.Label(frame_consume, text="Quantity:")
label_consume_quantity.grid(row=1, column=0, padx=10, pady=10)

entry_consume_quantity = tk.Entry(frame_consume)
entry_consume_quantity.grid(row=1, column=1, padx=10, pady=10)

label_consumer = tk.Label(frame_consume, text="Consumer:")
label_consumer.grid(row=2, column=0, padx=10, pady=10)

entry_consumer = tk.Entry(frame_consume)
entry_consumer.grid(row=2, column=1, padx=10, pady=10)

btn_consume = tk.Button(frame_consume, text="Consume Item", command=consumption)
btn_consume.grid(row=3, column=0, columnspan=2, pady=10)

# Run the app
root.mainloop()
